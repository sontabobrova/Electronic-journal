from datetime import date

import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.education.models import AcademicGroup, AcademicPeriod, StudentProfile, Subject, TeacherProfile, TeachingAssignment
from apps.journal.models import AttendanceRecord, ClassSession, Grade, GradeWork
from apps.reports.models import ReportRequest
from apps.users.models import UserRole


pytestmark = pytest.mark.django_db


def create_user(username: str, role: str):
    return get_user_model().objects.create_user(username=username, password="secret123", role=role)


def create_assignment(prefix: str):
    teacher_user = create_user(f"{prefix}-teacher", UserRole.TEACHER)
    teacher = TeacherProfile.objects.create(user=teacher_user, personnel_number=f"{prefix}-T-001")
    group = AcademicGroup.objects.create(name=f"{prefix}-743-2", enrollment_year=2023)
    subject = Subject.objects.create(name=f"{prefix} Безопасность", code=f"{prefix}-SEC-101")
    period = AcademicPeriod.objects.create(
        name=f"{prefix} Осенний семестр 2025",
        starts_at=date(2025, 9, 1),
        ends_at=date(2026, 1, 25),
    )
    assignment = TeachingAssignment.objects.create(teacher=teacher, subject=subject, group=group, period=period)
    return teacher_user, assignment


def create_student(username: str, student_id: int, group: AcademicGroup):
    user = create_user(username, UserRole.STUDENT)
    return StudentProfile.objects.create(user=user, student_id=student_id, group=group)


def build_report_data():
    admin = create_user("reports-admin", UserRole.ADMIN)
    teacher_user, assignment = create_assignment("reports-main")
    other_teacher_user, other_assignment = create_assignment("reports-other")
    student = create_student("reports-student", 9001, assignment.group)
    other_student = create_student("reports-other-student", 9002, other_assignment.group)

    work = GradeWork.objects.create(
        assignment=assignment,
        title="Основная контрольная",
        work_type="test",
        work_date=date(2025, 10, 1),
        max_score="5.00",
    )
    other_work = GradeWork.objects.create(
        assignment=other_assignment,
        title="Чужая контрольная",
        work_type="test",
        work_date=date(2025, 10, 1),
        max_score="5.00",
    )
    Grade.objects.create(work=work, student=student, value="5.00", created_by=teacher_user)
    Grade.objects.create(work=other_work, student=other_student, value="2.00", created_by=other_teacher_user)

    session = ClassSession.objects.create(assignment=assignment, session_date=date(2025, 10, 1), topic="Основное занятие")
    other_session = ClassSession.objects.create(
        assignment=other_assignment,
        session_date=date(2025, 10, 1),
        topic="Чужое занятие",
    )
    AttendanceRecord.objects.create(session=session, student=student, status="present", created_by=teacher_user)
    AttendanceRecord.objects.create(session=other_session, student=other_student, status="absent", created_by=other_teacher_user)

    return admin, teacher_user, student, other_student


def test_admin_generates_grades_csv_report(tmp_path, settings):
    settings.MEDIA_ROOT = tmp_path
    admin, _teacher_user, _student, _other_student = build_report_data()
    client = APIClient()
    client.force_authenticate(admin)

    response = client.post(
        reverse("report-requests-generate"),
        {"report_type": "grades", "file_format": "csv", "parameters": {}},
        format="json",
    )

    assert response.status_code == 201
    report = ReportRequest.objects.get()
    assert report.report_type == "grades"
    assert report.file_format == "csv"
    assert report.requested_by == admin
    assert report.file.name.endswith(".csv")
    content = report.file.read().decode("utf-8-sig")
    assert "Отчет по успеваемости" in content
    assert "Основная контрольная" in content
    assert "Чужая контрольная" in content
    assert response.data["download_url"].endswith(report.file.url)


def test_teacher_generates_only_own_xlsx_report(tmp_path, settings):
    settings.MEDIA_ROOT = tmp_path
    _admin, teacher_user, _student, _other_student = build_report_data()
    client = APIClient()
    client.force_authenticate(teacher_user)

    response = client.post(
        reverse("report-requests-generate"),
        {"report_type": "grades", "file_format": "xlsx", "parameters": {}},
        format="json",
    )

    assert response.status_code == 201
    report = ReportRequest.objects.get()
    workbook = load_workbook(report.file.path)
    worksheet = workbook.active
    values = [" ".join(str(cell or "") for cell in row) for row in worksheet.iter_rows(values_only=True)]
    joined = "\n".join(values)
    assert "Основная контрольная" in joined
    assert "Чужая контрольная" not in joined


def test_student_cannot_generate_report_for_another_student(tmp_path, settings):
    settings.MEDIA_ROOT = tmp_path
    _admin, _teacher_user, student, other_student = build_report_data()
    client = APIClient()
    client.force_authenticate(student.user)

    response = client.post(
        reverse("report-requests-generate"),
        {"report_type": "attendance", "file_format": "csv", "parameters": {"student": other_student.id}},
        format="json",
    )

    assert response.status_code == 400
    assert ReportRequest.objects.count() == 0


def test_student_generates_own_attendance_pdf_report(tmp_path, settings):
    settings.MEDIA_ROOT = tmp_path
    _admin, _teacher_user, student, _other_student = build_report_data()
    client = APIClient()
    client.force_authenticate(student.user)

    response = client.post(
        reverse("report-requests-generate"),
        {"report_type": "attendance", "file_format": "pdf", "parameters": {}},
        format="json",
    )

    assert response.status_code == 201
    report = ReportRequest.objects.get()
    assert report.file.name.endswith(".pdf")
    assert report.file.size > 0


def test_report_history_is_private_for_non_admin(tmp_path, settings):
    settings.MEDIA_ROOT = tmp_path
    admin, teacher_user, student, _other_student = build_report_data()
    ReportRequest.objects.create(requested_by=admin, report_type="grades", file_format="csv", parameters={})
    ReportRequest.objects.create(requested_by=teacher_user, report_type="grades", file_format="csv", parameters={})
    ReportRequest.objects.create(requested_by=student.user, report_type="attendance", file_format="csv", parameters={})

    client = APIClient()
    client.force_authenticate(teacher_user)
    teacher_response = client.get(reverse("report-requests-list"))
    assert teacher_response.status_code == 200
    assert len(teacher_response.data) == 1

    client.force_authenticate(admin)
    admin_response = client.get(reverse("report-requests-list"))
    assert admin_response.status_code == 200
    assert len(admin_response.data) == 3
